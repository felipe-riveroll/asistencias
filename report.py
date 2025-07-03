import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import messagebox


def _get_expected_seconds_for_day(row_data, expected_hours_df, cache):
    if expected_hours_df is None:
        return 0.0
    try:
        if row_data.get("Turno") == "Totales":
            return 0.0

        if not pd.notnull(row_data.get("Fecha")):
            return 0.0

        dia_semana_str = row_data.get("Día", "")
        if not dia_semana_str or dia_semana_str not in expected_hours_df.columns:
            return 0.0

        emp_id_str = str(row_data.get("ID Empleado", "")).strip()
        if not emp_id_str:
            return 0.0

        cache_key = f"{emp_id_str}_{dia_semana_str}"
        if cache_key in cache:
            return float(cache[cache_key])

        try:
            employee_id_num = int(float(emp_id_str))
        except ValueError:
            return 0.0

        emp_mask = expected_hours_df["Employee"] == employee_id_num
        if not emp_mask.any():
            cache[cache_key] = 0.0
            return 0.0

        value_seconds = expected_hours_df.loc[emp_mask, dia_semana_str].iloc[0]
        result_seconds = float(value_seconds) if pd.notnull(value_seconds) else 0.0

        cache[cache_key] = result_seconds
        return result_seconds
    except Exception:
        return 0.0


def generate_report(src, dst, expected_hours_df=None, expected_hours_cache=None):
    if expected_hours_cache is None:
        expected_hours_cache = {}

    df_excel = pd.read_excel(src)
    if {"Employee Name", "Time"}.difference(df_excel.columns):
        raise ValueError(
            "Las columnas requeridas 'Employee Name' y 'Time' no se encontraron."
        )

    df_proc = df_excel.copy()
    df_proc["Time"] = pd.to_datetime(df_proc["Time"], errors="coerce")
    df_proc.dropna(subset=["Time"], inplace=True)

    df_proc["Day_raw"] = df_proc["Time"].dt.date
    df_proc["WorkDay"] = df_proc.apply(
        lambda r: r["Day_raw"] - datetime.timedelta(days=1)
        if r["Time"].hour < 6
        else r["Day_raw"],
        axis=1,
    )
    if "Shift" not in df_proc.columns:
        df_proc["Shift"] = ""
    df_proc["Shift"] = df_proc["Shift"].fillna("")

    df_turno = df_proc[df_proc["Shift"] != ""].copy()
    df_sin_turno = df_proc[df_proc["Shift"] == ""].copy()
    if not df_sin_turno.empty:
        df_sin_turno.loc[:, "Merged"] = False

    grouped = (
        df_turno.groupby(["Employee Name", "Shift", "WorkDay"])
        .agg(
            checadas_list=(
                "Time",
                lambda ts: sorted([pd.to_datetime(t) for t in ts if pd.notnull(t)]),
            )
        )
        .reset_index()
    )

    def merge_no_shift(gdf, ns_df):
        res_df = gdf.copy()
        if "checadas_list" not in res_df.columns:
            res_df["checadas_list"] = [[] for _ in range(len(res_df))]
        else:
            res_df["checadas_list"] = res_df["checadas_list"].apply(
                lambda x: list(x) if isinstance(x, (list, pd.Series)) else []
            )

        if ns_df.empty:
            return res_df
        for idx, row_ns in ns_df.iterrows():
            time_to_add = pd.to_datetime(row_ns["Time"])
            mask = (res_df["Employee Name"] == row_ns["Employee Name"]) & (
                res_df["WorkDay"] == row_ns["WorkDay"]
            )
            if mask.any():
                res_idx = res_df[mask].index[0]
                current_list = res_df.at[res_idx, "checadas_list"]
                if time_to_add not in current_list:
                    res_df.at[res_idx, "checadas_list"] = sorted(current_list + [time_to_add])
                ns_df.loc[idx, "Merged"] = True
            else:
                ns_df.loc[idx, "Merged"] = False
        extra_df = ns_df[~ns_df["Merged"]]
        if not extra_df.empty:
            add_df = (
                extra_df.groupby(["Employee Name", "WorkDay"])
                .agg(
                    checadas_list=(
                        "Time",
                        lambda ts: sorted([pd.to_datetime(t) for t in ts if pd.notnull(t)]),
                    )
                )
                .reset_index()
            )
            add_df["Shift"] = ""
            res_df = pd.concat([res_df, add_df], ignore_index=True)
        return res_df

    grouped = merge_no_shift(grouped, df_sin_turno)
    grouped.rename(columns={"WorkDay": "Fecha_raw"}, inplace=True)
    grouped.sort_values(["Employee Name", "Fecha_raw"], inplace=True)

    def calc_actual_worked_hours(lst_times):
        if not lst_times or len(lst_times) < 2:
            return pd.Timedelta(0)
        return lst_times[-1] - lst_times[0]

    grouped["total_timedelta_actual"] = grouped["checadas_list"].apply(calc_actual_worked_hours)
    fmt_timedelta_to_str = (
        lambda td: f"{int(td.total_seconds()//3600):02d}:{int(td.total_seconds()%3600//60):02d}:{int(round(td.total_seconds()%60)):02d}"
        if pd.notnull(td) and td.total_seconds() > 0
        else "00:00:00"
    )
    grouped["Horas totales_str"] = grouped["total_timedelta_actual"].apply(fmt_timedelta_to_str)

    grouped["Checadas_str_list"] = grouped["checadas_list"].apply(
        lambda ts: [t.strftime("%H:%M:%S") for t in ts if pd.notnull(t)]
    )
    max_chec = grouped["Checadas_str_list"].str.len().max()
    if pd.isna(max_chec) or max_chec == 0:
        max_chec = 1

    chec_df_data = {}
    for i in range(int(max_chec)):
        chec_df_data[f"Checada {i+1}"] = grouped["Checadas_str_list"].apply(
            lambda x: x[i] if i < len(x) else None
        )
    chec_df = pd.DataFrame(chec_df_data, index=grouped.index)

    if "Employee" in df_excel.columns and "Employee Name" in df_excel.columns:
        id_map = (
            df_excel[["Employee Name", "Employee"]]
            .drop_duplicates("Employee Name", keep="first")
            .set_index("Employee Name")["Employee"]
            .to_dict()
        )
        grouped["ID Empleado_val"] = grouped["Employee Name"].map(id_map).fillna("")
    else:
        grouped["ID Empleado_val"] = ""

    report_cols_from_grouped = [
        "ID Empleado_val",
        "Employee Name",
        "Shift",
        "Fecha_raw",
        "Horas totales_str",
    ]
    for col_name in report_cols_from_grouped:
        if col_name not in grouped:
            grouped[col_name] = None if col_name not in ["Shift", "ID Empleado_val"] else ""

    report_df = pd.concat([grouped[report_cols_from_grouped], chec_df], axis=1)
    report_df.rename(
        columns={
            "ID Empleado_val": "ID Empleado",
            "Employee Name": "Nombre del empleado",
            "Shift": "Turno",
            "Fecha_raw": "Fecha",
            "Horas totales_str": "Horas totales",
        },
        inplace=True,
    )

    report_df["Fecha"] = pd.to_datetime(report_df["Fecha"], errors="coerce").dt.date

    dias_semana = {0: "Lunes", 1: "Martes", 2: "Miércoles", 3: "Jueves", 4: "Viernes", 5: "Sábado", 6: "Domingo"}
    report_df["Día"] = pd.to_datetime(report_df["Fecha"]).dt.weekday.map(dias_semana).fillna("")

    report_df["Horas esperadas"] = report_df.apply(
        lambda row: _get_expected_seconds_for_day(row, expected_hours_df, expected_hours_cache),
        axis=1,
    )

    core_cols = [
        "ID Empleado",
        "Nombre del empleado",
        "Turno",
        "Fecha",
        "Día",
        "Horas esperadas",
        "Horas totales",
    ]
    checada_cols_in_report = sorted(
        [col for col in report_df.columns if col.startswith("Checada ")],
        key=lambda x: int(x.split(" ")[1]),
    )
    final_report_columns_ordered = core_cols + checada_cols_in_report

    display_report_df = report_df.copy()
    for col_name in final_report_columns_ordered:
        if col_name not in display_report_df.columns:
            display_report_df[col_name] = None
    display_report_df = display_report_df[final_report_columns_ordered]

    summary_actual = (
        grouped.groupby(["ID Empleado_val", "Employee Name"])
        .agg(
            total_worked_time_sum=("total_timedelta_actual", "sum"),
            first_day_worked=("Fecha_raw", "min"),
            last_day_worked=("Fecha_raw", "max"),
            days_actually_worked=("Fecha_raw", "nunique"),
        )
        .reset_index()
    )
    summary_actual.rename(
        columns={"ID Empleado_val": "ID Empleado", "Employee Name": "Nombre"}, inplace=True
    )

    if not report_df.empty and "Horas esperadas" in report_df.columns:
        total_expected_summary = (
            report_df.groupby(["ID Empleado", "Nombre del empleado"])["Horas esperadas"]
            .sum()
            .reset_index(name="Total Segundos Esperados")
        )
        total_expected_summary.rename(columns={"Nombre del empleado": "Nombre"}, inplace=True)
        resumen_df = pd.merge(summary_actual, total_expected_summary, on=["ID Empleado", "Nombre"], how="left")
        resumen_df["Total Segundos Esperados"] = resumen_df["Total Segundos Esperados"].fillna(0)
    else:
        resumen_df = summary_actual.copy()
        resumen_df["Total Segundos Esperados"] = 0.0

    resumen_df["Días del periodo"] = (
        resumen_df["last_day_worked"] - resumen_df["first_day_worked"]
    ).apply(lambda td: td.days + 1 if pd.notnull(td) else 0)
    resumen_df["Horas trabajadas"] = resumen_df["total_worked_time_sum"].apply(fmt_timedelta_to_str)
    resumen_df["Horas Trabajadas (Segundos)"] = resumen_df["total_worked_time_sum"].apply(
        lambda td: td.total_seconds() if pd.notnull(td) else 0
    )
    resumen_df["Total Segundos Esperados"] = pd.to_numeric(
        resumen_df["Total Segundos Esperados"], errors="coerce"
    ).fillna(0)
    resumen_df["Diferencia (Segundos)"] = (
        resumen_df["Horas Trabajadas (Segundos)"] - resumen_df["Total Segundos Esperados"]
    )

    def format_seconds_to_hhmmss_with_sign(total_seconds):
        if pd.isna(total_seconds):
            return "00:00:00"
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        h = int(total_seconds // 3600)
        m = int((total_seconds % 3600) // 60)
        s = int(round(total_seconds % 60))
        return f"{sign}{h:02d}:{m:02d}:{s:02d}"

    resumen_df["Diferencia (HH:MM:SS)"] = resumen_df["Diferencia (Segundos)"].apply(
        format_seconds_to_hhmmss_with_sign
    )

    resumen_df.rename(columns={"days_actually_worked": "Días trabajados"}, inplace=True, errors="ignore")

    resumen_df_cols_final = [
        "ID Empleado",
        "Nombre",
        "Días del periodo",
        "Días trabajados",
        "Horas trabajadas",
        "Horas Trabajadas (Segundos)",
        "Total Segundos Esperados",
        "Diferencia (Segundos)",
        "Diferencia (HH:MM:SS)",
    ]
    for col in resumen_df_cols_final:
        if col not in resumen_df.columns:
            default_val = ""
            if col in [
                "Días del periodo",
                "Días trabajados",
                "Horas Trabajadas (Segundos)",
                "Total Segundos Esperados",
                "Diferencia (Segundos)",
            ]:
                default_val = 0.0
            resumen_df[col] = default_val

    resumen_df = resumen_df[resumen_df_cols_final]

    total_rows_for_detail_list = []
    for _, r_resumen_row in resumen_df.iterrows():
        emp_name_for_total = r_resumen_row["Nombre"]
        sum_numeric_expected_seconds_for_emp = r_resumen_row["Total Segundos Esperados"]
        dias_trabajados_val = (
            int(r_resumen_row["Días trabajados"]) if pd.notnull(r_resumen_row["Días trabajados"]) else 0
        )
        total_row_dict = {
            "ID Empleado": r_resumen_row.get("ID Empleado", ""),
            "Nombre del empleado": emp_name_for_total,
            "Turno": "Totales",
            "Fecha": dias_trabajados_val,
            "Día": "",
            "Horas esperadas": sum_numeric_expected_seconds_for_emp,
            "Horas totales": r_resumen_row["Horas trabajadas"],
        }
        for c_col in checada_cols_in_report:
            total_row_dict[c_col] = ""
        total_rows_for_detail_list.append(total_row_dict)

    totals_to_append_df = pd.DataFrame(total_rows_for_detail_list)
    if not totals_to_append_df.empty:
        totals_to_append_df = totals_to_append_df.reindex(columns=display_report_df.columns)

    final_detail_sheet_dfs = []
    if not display_report_df.empty:
        for emp_name, daily_data_group in display_report_df.groupby("Nombre del empleado", sort=False):
            final_detail_sheet_dfs.append(daily_data_group.sort_values("Fecha"))
            emp_total_row = totals_to_append_df[totals_to_append_df["Nombre del empleado"] == emp_name]
            if not emp_total_row.empty:
                final_detail_sheet_dfs.append(emp_total_row)

    if not final_detail_sheet_dfs:
        final_detail_report_df = pd.DataFrame(columns=final_report_columns_ordered)
    else:
        final_detail_report_df = pd.concat(final_detail_sheet_dfs, ignore_index=True)

    if "Fecha" in final_detail_report_df.columns:
        def format_fecha_col(val):
            if isinstance(val, (datetime.date, pd.Timestamp)):
                return val.strftime("%Y-%m-%d")
            return val

        final_detail_report_df["Fecha"] = final_detail_report_df["Fecha"].apply(format_fecha_col)

    with pd.ExcelWriter(dst, engine="openpyxl") as writer:
        final_detail_report_df.to_excel(writer, index=False, sheet_name="Detalle")
        resumen_df.to_excel(writer, index=False, sheet_name="Resumen")

    return resumen_df


def format_excel(path, resumen_data_df=None):
    wb = load_workbook(path)

    yellow_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    dark_orange_fill = PatternFill(start_color="1CC0EE", end_color="1CC0EE", fill_type="solid")

    def _format_ws(ws, is_resumen_sheet=False, df_data_for_resumen=None):
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        bold_font = Font(bold=True)
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if ws.max_row == 0:
            return

        col_names_map = {cell.value: get_column_letter(cell.column) for cell in ws[1]}

        for c_idx_plus_1 in range(1, ws.max_column + 1):
            cell = ws.cell(1, c_idx_plus_1)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin

        if ws.title == "Detalle":
            turno_col_letter = col_names_map.get("Turno")
            if turno_col_letter:
                for r_idx_plus_1 in range(2, ws.max_row + 1):
                    if ws[f"{turno_col_letter}{r_idx_plus_1}"].value == "Totales":
                        for c_idx_plus_1_total in range(1, ws.max_column + 1):
                            cell_total = ws.cell(r_idx_plus_1, c_idx_plus_1_total)
                            cell_total.fill = total_fill
                            cell_total.font = bold_font
                            cell_total.border = thin

        if is_resumen_sheet and df_data_for_resumen is not None:
            diferencia_hhmmss_col_letter = col_names_map.get("Diferencia (HH:MM:SS)")
            diferencia_segundos_col_name = "Diferencia (Segundos)"

            if (
                diferencia_hhmmss_col_letter
                and diferencia_segundos_col_name in df_data_for_resumen.columns
            ):
                for r_idx, (_, row_data) in enumerate(df_data_for_resumen.iterrows()):
                    excel_row_num = r_idx + 2
                    cell_to_format = ws[f"{diferencia_hhmmss_col_letter}{excel_row_num}"]
                    diferencia_sec_valor = row_data[diferencia_segundos_col_name]

                    if pd.notnull(diferencia_sec_valor):
                        if diferencia_sec_valor < 0:
                            cell_to_format.fill = yellow_fill
                        elif diferencia_sec_valor > 0:
                            cell_to_format.fill = dark_orange_fill

        for col_letter_obj in ws.columns:
            column_letter_str = col_letter_obj[0].column_letter
            max_len = 0
            for cell_in_col in col_letter_obj:
                if cell_in_col.value is not None:
                    try:
                        max_len = max(max_len, len(str(cell_in_col.value)))
                    except Exception:
                        pass

            adjusted_width = max_len + 3
            header_value = ws[f"{column_letter_str}1"].value

            min_widths = {
                "ID Empleado": 12,
                "Nombre del empleado": 30,
                "Nombre": 30,
                "Turno": 10,
                "Fecha": 12,
                "Día": 12,
                "Horas esperadas": 20,
                "Horas totales": 15,
                "Horas trabajadas": 15,
                "Horas Trabajadas (Segundos)": 22,
                "Total Segundos Esperados": 22,
                "Diferencia (Segundos)": 22,
                "Diferencia (HH:MM:SS)": 22,
                "Días del periodo": 18,
                "Días trabajados": 18,
            }

            default_min_width = 10 if str(header_value).startswith("Checada") else 12
            adjusted_width = max(adjusted_width, min_widths.get(header_value, default_min_width))
            ws.column_dimensions[column_letter_str].width = adjusted_width

    for sheet_name_iter in wb.sheetnames:
        current_ws = wb[sheet_name_iter]
        if sheet_name_iter == "Resumen":
            _format_ws(current_ws, is_resumen_sheet=True, df_data_for_resumen=resumen_data_df)
        else:
            _format_ws(current_ws)
    try:
        wb.save(path)
    except Exception as e_save:  # pragma: no cover - relies on Excel
        messagebox.showerror(
            "Error al guardar",
            f"No se pudo guardar el archivo Excel:\n{e_save}\n\nAsegúrese de que el archivo no esté abierto.",
        )
        print(f"Error al guardar Excel: {e_save}")

