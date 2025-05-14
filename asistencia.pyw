import pandas as pd
import os
import datetime
import subprocess
import traceback # Keep for debugging if needed
from tkinter import Tk, Label, Button, Entry, StringVar, filedialog, Frame, ttk, messagebox, Toplevel
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment 
from openpyxl.utils import get_column_letter 
# from scipy.stats import wilcoxon # Eliminada

pd.options.mode.chained_assignment = None 

class CheckadorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Procesador de Checadas")
        self.root.geometry("750x520")
        self.root.resizable(True, True)

        self.primary_color   = "#2c3e50"
        self.secondary_color = "#3498db"
        self.bg_color        = "#f5f5f5"
        self.text_color      = "#333333"
        self.success_color   = "#27ae60"
        self.warning_color   = "#f39c12" 
        self.error_color     = "#e74c3c"
        self.orange_dark_color = "#E67E22" 

        self.root.configure(bg=self.bg_color)
        self.expected_hours_df = self._load_expected_hours_data()
        self.expected_hours_cache = {}

        self.status_frame = Frame(root, bg="#e0e0e0", relief="ridge", bd=1)
        self.status_frame.pack(side="bottom", fill="x")
        self.status_label = Label(self.status_frame, text="Listo para procesar", font=("Segoe UI", 10),
                                  bg="#e0e0e0", fg=self.primary_color, padx=10, pady=8)
        self.status_label.pack(fill="x")

        style = ttk.Style()
        style.configure("TButton", font=("Segoe UI", 10))
        style.configure("TEntry",  font=("Segoe UI", 10))
        style.configure("TLabel",  font=("Segoe UI", 10), background=self.bg_color)

        self.input_file_path = StringVar()
        self.output_file_name = StringVar(value=f"reporte_checador_{datetime.datetime.now().strftime('%d%m%Y')}")

        main = Frame(root, bg=self.bg_color, padx=30, pady=20); main.pack(fill="both", expand=True)

        try:
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logo_asia.png")
            if os.path.exists(logo_path):
                img = Image.open(logo_path); w, h = img.size; nw = 200; nh = int(nw / w * h)
                logo = ImageTk.PhotoImage(img.resize((nw, nh), Image.LANCZOS))
                Label(main, image=logo, bg=self.bg_color).pack(pady=(0, 15)); self._logo_ref = logo
        except Exception as e_logo:
            print(f"Error loading logo: {e_logo}")

        banner = Frame(main, bg=self.primary_color, height=60); banner.pack(fill="x", pady=(0, 20))
        Label(banner, text="Procesador de Checadas de Empleados", font=("Segoe UI", 18, "bold"), bg=self.primary_color, fg="white").pack(pady=10)

        form = Frame(main, bg=self.bg_color, padx=20, pady=10); form.pack(fill="both", expand=True)
        row1 = Frame(form, bg=self.bg_color, pady=10); row1.pack(fill="x")
        Label(row1, text="Archivo Excel de Checadas:", font=("Segoe UI", 11), bg=self.bg_color, fg=self.text_color).pack(side="left", padx=(0, 10))
        Entry(row1, textvariable=self.input_file_path, font=("Segoe UI", 10), bd=1, relief="solid").pack(side="left", fill="x", expand=True, ipady=3)
        Button(row1, text="Examinar...", command=self.browse_file, font=("Segoe UI", 10), bg=self.secondary_color, fg="white", relief="flat").pack(side="left", padx=(10, 0))

        row2 = Frame(form, bg=self.bg_color, pady=10); row2.pack(fill="x")
        Label(row2, text="Nombre del archivo de salida:", font=("Segoe UI", 11), bg=self.bg_color, fg=self.text_color).pack(side="left", padx=(0, 10))
        Entry(row2, textvariable=self.output_file_name, font=("Segoe UI", 10), bd=1, relief="solid").pack(side="left", fill="x", expand=True, ipady=3)
        Label(row2, text=".xlsx", font=("Segoe UI", 11), bg=self.bg_color, fg=self.text_color).pack(side="left")

        ttk.Separator(form, orient="horizontal").pack(fill="x", pady=20)

        actions = Frame(form, bg=self.bg_color, pady=20); actions.pack(fill="x")
        self.process_button = Button(actions, text="Procesar Archivo", command=self.generate_report, font=("Segoe UI", 12, "bold"), bg=self.secondary_color, fg="white", relief="flat", padx=20, pady=8)
        self.process_button.pack(pady=10)
        self.progress = ttk.Progressbar(actions, orient="horizontal", length=500, mode="indeterminate")

    def _load_expected_hours_data(self):
        try:
            base_path = os.path.dirname(os.path.abspath(__file__))
            expected_hours_path = os.path.join(base_path, 'expected_hours_data.csv')
            if os.path.exists(expected_hours_path):
                with open(expected_hours_path, 'r', encoding='utf-8') as f:
                    first_line = f.readline().strip()
                    skip_rows = 1 if first_line.startswith('//') else 0
                df_expected = pd.read_csv(expected_hours_path, skiprows=skip_rows)
                required_cols = ['Employee', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
                if all(col in df_expected.columns for col in required_cols):
                    df_expected['Employee'] = pd.to_numeric(df_expected['Employee'], errors='coerce').fillna(0).astype(int)
                    for day_col in required_cols[1:]: 
                        if day_col in df_expected.columns:
                            df_expected[day_col] = pd.to_numeric(df_expected[day_col], errors='coerce').fillna(0)
                    return df_expected
                else:
                    print(f"Advertencia: Faltan columnas en 'expected_hours_data.csv'. Esperadas: {required_cols}.")
            else:
                messagebox.showwarning("Archivo no encontrado", f"No se encontró 'expected_hours_data.csv'.")
        except Exception as e:
            messagebox.showerror("Error al cargar horas", f"Error cargando 'expected_hours_data.csv': {e}")
            print(f"Error cargando 'expected_hours_data.csv': {e}\n{traceback.format_exc()}")
        return None

    def browse_file(self):
        fp = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx;*.xls"), ("Todos", "*.*")])
        if fp:
            self.input_file_path.set(fp)
            self.status_label.configure(text=f"Archivo seleccionado: {os.path.basename(fp)}", fg=self.primary_color, bg="#e0e0e0")

    def _set_status(self, msg, kind="info"):
        cfg = {"success":("white", self.success_color), "warning":("white", self.warning_color), "error":("white", self.error_color), "info":(self.primary_color, "#e0e0e0")}
        fg, bg = cfg.get(kind, cfg["info"]); self.status_label.configure(text=msg, fg=fg, bg=bg); self.root.update()

    def _toggle_busy(self, busy: bool):
        if busy:
            self.process_button.configure(state="disabled", text="Procesando...", bg="#95a5a6")
            self.progress.pack(pady=10); self.progress.start(10)
        else:
            self.process_button.configure(state="normal", text="Procesar Archivo", bg=self.secondary_color)
            self.progress.stop(); self.progress.pack_forget()

    def _show_success_dialog(self, path: str):
        dlg = Toplevel(self.root); dlg.title("Proceso completado"); dlg.geometry("450x250"); dlg.resizable(False, False); dlg.configure(bg="white")
        content = Frame(dlg, bg="white", padx=20, pady=10); content.pack(fill="both", expand=True)
        try:
            logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Logo_asia.png")
            if os.path.exists(logo_path):
                img = Image.open(logo_path); w, h = img.size; nw = 150; nh = int(nw / w * h)
                logo = ImageTk.PhotoImage(img.resize((nw, nh), Image.LANCZOS)); Label(content, image=logo, bg="white").pack(pady=(0,10)); dlg._logo=logo
            else:
                 Label(content, text="ⓘ", font=("Segoe UI", 24), fg=self.secondary_color, bg="white").pack(pady=(10,5))
        except Exception:
            Label(content, text="ⓘ", font=("Segoe UI", 24), fg=self.secondary_color, bg="white").pack(pady=(10,5))
        Label(content, text="El reporte se ha generado correctamente.", font=("Segoe UI", 11), bg="white").pack(pady=5)
        short = path if len(path)<=50 else os.path.join(os.path.dirname(path)[:10]+"...", os.path.basename(path))
        pf = Frame(content, bg="white"); pf.pack(pady=5, fill="x")
        Label(pf, text="Ubicación:", font=("Segoe UI", 10), bg="white").pack(side="left")
        Label(pf, text=short, font=("Segoe UI", 9), fg="#555", bg="white").pack(side="left")
        bf = Frame(content, bg="white"); bf.pack(fill="x", pady=10)
        def _open_file_action():
            try:
                if os.name=='nt': os.startfile(path)
                elif os.name=='posix': subprocess.call(('open' if os.path.exists('/usr/bin/open') else 'xdg-open', path))
                dlg.destroy()
            except Exception as e_open:
                messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e_open}")
        Button(bf, text="Abrir archivo", font=("Segoe UI", 10), bg=self.secondary_color, fg="white", width=12, command=_open_file_action).pack(side="right", padx=5)
        Button(bf, text="Aceptar", font=("Segoe UI", 10), bg="#f0f0f0", width=10, command=dlg.destroy).pack(side="right", padx=5)
        dlg.transient(self.root); dlg.grab_set(); dlg.protocol("WM_DELETE_WINDOW", dlg.destroy)
        dlg.update_idletasks(); w_dlg, h_dlg = dlg.winfo_width(), dlg.winfo_height(); x = (dlg.winfo_screenwidth()-w_dlg)//2; y = (dlg.winfo_screenheight()-h_dlg)//2; dlg.geometry(f"{w_dlg}x{h_dlg}+{x}+{y}")

    def generate_report(self):
        src = self.input_file_path.get().strip()
        if not src:
            self._set_status("Seleccione un archivo", "error"); messagebox.showerror("Error", "Debe seleccionar un archivo de entrada"); return
        
        output_filename = self.output_file_name.get().strip() + ".xlsx"
        if os.path.isabs(src) and os.path.isdir(os.path.dirname(src)):
            dst_folder = os.path.dirname(src)
        else: 
            dst_folder = os.path.dirname(os.path.abspath(__file__))
        dst = os.path.join(dst_folder, output_filename)

        try:
            self._toggle_busy(True); self._set_status("Procesando archivo...", "info")
            df_excel = pd.read_excel(src) 
            if {'Employee Name','Time'}.difference(df_excel.columns):
                raise ValueError("Las columnas requeridas 'Employee Name' y 'Time' no se encontraron.")
            
            df_proc = df_excel.copy()
            df_proc['Time'] = pd.to_datetime(df_proc['Time'], errors='coerce')
            df_proc.dropna(subset=['Time'], inplace=True)

            df_proc['Day_raw'] = df_proc['Time'].dt.date
            df_proc['WorkDay'] = df_proc.apply(lambda r: r['Day_raw'] - datetime.timedelta(days=1) if r['Time'].hour < 6 else r['Day_raw'], axis=1)
            if 'Shift' not in df_proc.columns: df_proc['Shift'] = ''
            df_proc['Shift'] = df_proc['Shift'].fillna('')

            df_turno = df_proc[df_proc['Shift'] != ''].copy()
            df_sin_turno = df_proc[df_proc['Shift'] == ''].copy()
            if not df_sin_turno.empty:
                df_sin_turno.loc[:, 'Merged'] = False

            grouped = (df_turno.groupby(['Employee Name', 'Shift', 'WorkDay'])
                       .agg(checadas_list=('Time', lambda ts: sorted([pd.to_datetime(t) for t in ts if pd.notnull(t)])))
                       .reset_index())

            def merge_no_shift(gdf, ns_df):
                res_df = gdf.copy()
                if 'checadas_list' not in res_df.columns:
                    res_df['checadas_list'] = [[] for _ in range(len(res_df))]
                else:
                    res_df['checadas_list'] = res_df['checadas_list'].apply(lambda x: list(x) if isinstance(x, (list, pd.Series)) else [])

                if ns_df.empty: return res_df
                for idx, row_ns in ns_df.iterrows():
                    time_to_add = pd.to_datetime(row_ns['Time'])
                    mask = (res_df['Employee Name'] == row_ns['Employee Name']) & (res_df['WorkDay'] == row_ns['WorkDay'])
                    if mask.any():
                        res_idx = res_df[mask].index[0]
                        current_list = res_df.at[res_idx, 'checadas_list']
                        if time_to_add not in current_list:
                            res_df.at[res_idx, 'checadas_list'] = sorted(current_list + [time_to_add])
                        ns_df.loc[idx, 'Merged'] = True
                    else:
                        ns_df.loc[idx, 'Merged'] = False
                extra_df = ns_df[~ns_df['Merged']]
                if not extra_df.empty:
                    add_df = (extra_df.groupby(['Employee Name', 'WorkDay'])
                              .agg(checadas_list=('Time', lambda ts: sorted([pd.to_datetime(t) for t in ts if pd.notnull(t)])))
                              .reset_index())
                    add_df['Shift'] = ''
                    res_df = pd.concat([res_df, add_df], ignore_index=True)
                return res_df

            grouped = merge_no_shift(grouped, df_sin_turno)
            grouped.rename(columns={'WorkDay': 'Fecha_raw'}, inplace=True)
            grouped.sort_values(['Employee Name', 'Fecha_raw'], inplace=True)

            def calc_actual_worked_hours(lst_times):
                if not lst_times or len(lst_times) < 2: return pd.Timedelta(0)
                return lst_times[-1] - lst_times[0]

            grouped['total_timedelta_actual'] = grouped['checadas_list'].apply(calc_actual_worked_hours)
            fmt_timedelta_to_str = lambda td: f"{int(td.total_seconds()//3600):02d}:{int(td.total_seconds()%3600//60):02d}:{int(round(td.total_seconds()%60)):02d}" if pd.notnull(td) and td.total_seconds() > 0 else "00:00:00"
            grouped['Horas totales_str'] = grouped['total_timedelta_actual'].apply(fmt_timedelta_to_str)
            
            grouped['Checadas_str_list'] = grouped['checadas_list'].apply(lambda ts: [t.strftime('%H:%M:%S') for t in ts if pd.notnull(t)])
            max_chec = grouped['Checadas_str_list'].str.len().max()
            if pd.isna(max_chec) or max_chec == 0: max_chec = 1
            
            chec_df_data = {}
            for i in range(int(max_chec)):
                chec_df_data[f'Checada {i+1}'] = grouped['Checadas_str_list'].apply(lambda x: x[i] if i < len(x) else None)
            chec_df = pd.DataFrame(chec_df_data, index=grouped.index)

            if 'Employee' in df_excel.columns and 'Employee Name' in df_excel.columns:
                id_map = df_excel[['Employee Name', 'Employee']].drop_duplicates('Employee Name', keep='first').set_index('Employee Name')['Employee'].to_dict()
                grouped['ID Empleado_val'] = grouped['Employee Name'].map(id_map).fillna("") 
            else:
                grouped['ID Empleado_val'] = "" 

            report_cols_from_grouped = ['ID Empleado_val', 'Employee Name', 'Shift', 'Fecha_raw', 'Horas totales_str']
            for col_name in report_cols_from_grouped:
                if col_name not in grouped:
                    grouped[col_name] = None if col_name not in ['Shift', 'ID Empleado_val'] else ''
            
            report_df = pd.concat([grouped[report_cols_from_grouped], chec_df], axis=1)
            report_df.rename(columns={'ID Empleado_val': 'ID Empleado',
                                   'Employee Name': 'Nombre del empleado',
                                   'Shift': 'Turno',
                                   'Fecha_raw': 'Fecha', # Fecha_raw (WorkDay) is now 'Fecha'
                                   'Horas totales_str': 'Horas totales'}, inplace=True)

            # Ensure 'Fecha' is datetime.date objects for display, not full timestamps
            report_df['Fecha'] = pd.to_datetime(report_df['Fecha'], errors='coerce').dt.date


            dias_semana = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}
            # Apply weekday to the 'Fecha' column (which should be date objects)
            report_df['Día'] = pd.to_datetime(report_df['Fecha']).dt.weekday.map(dias_semana).fillna('')


            def get_expected_seconds_for_day(row_data):
                if self.expected_hours_df is None: return 0.0
                try:
                    if row_data.get('Turno') == 'Totales': return 0.0 # Skip for total rows being built
                    
                    # 'Fecha' should be a date object here, 'Día' is the Spanish day name
                    if not pd.notnull(row_data.get('Fecha')): return 0.0
                    
                    dia_semana_str = row_data.get('Día', '') # This is already the Spanish day name
                    if not dia_semana_str or dia_semana_str not in self.expected_hours_df.columns: return 0.0
                    
                    emp_id_str = str(row_data.get('ID Empleado', "")).strip() 
                    if not emp_id_str: return 0.0 
                    
                    cache_key = f"{emp_id_str}_{dia_semana_str}"
                    if cache_key in self.expected_hours_cache: return float(self.expected_hours_cache[cache_key])
                    
                    try: employee_id_num = int(float(emp_id_str)) 
                    except ValueError: return 0.0
                        
                    emp_mask = self.expected_hours_df['Employee'] == employee_id_num
                    if not emp_mask.any():
                        self.expected_hours_cache[cache_key] = 0.0; return 0.0
                    
                    value_seconds = self.expected_hours_df.loc[emp_mask, dia_semana_str].iloc[0]
                    result_seconds = float(value_seconds) if pd.notnull(value_seconds) else 0.0
                    
                    self.expected_hours_cache[cache_key] = result_seconds
                    return result_seconds
                except Exception: return 0.0
            
            report_df['Horas esperadas'] = report_df.apply(get_expected_seconds_for_day, axis=1)
            
            core_cols = ['ID Empleado', 'Nombre del empleado', 'Turno', 'Fecha', 'Día', 'Horas esperadas', 'Horas totales']
            checada_cols_in_report = sorted([col for col in report_df.columns if col.startswith('Checada ')], 
                                            key=lambda x: int(x.split(' ')[1]))
            final_report_columns_ordered = core_cols + checada_cols_in_report
            
            display_report_df = report_df.copy() 
            for col_name in final_report_columns_ordered:
                if col_name not in display_report_df.columns: display_report_df[col_name] = None 
            display_report_df = display_report_df[final_report_columns_ordered]

            summary_actual = (grouped.groupby(['ID Empleado_val', 'Employee Name'])
                               .agg(total_worked_time_sum=('total_timedelta_actual', 'sum'),
                                    first_day_worked=('Fecha_raw', 'min'),
                                    last_day_worked=('Fecha_raw', 'max'),
                                    days_actually_worked=('Fecha_raw', 'nunique'))
                               .reset_index())
            summary_actual.rename(columns={'ID Empleado_val': 'ID Empleado', 'Employee Name': 'Nombre'}, inplace=True)

            if not report_df.empty and 'Horas esperadas' in report_df.columns:
                # Use report_df for summing total expected seconds as it has ID Empleado and Nombre del empleado
                # Group by the same columns used in summary_actual after its rename
                total_expected_summary = (report_df.groupby(['ID Empleado', 'Nombre del empleado'])['Horas esperadas']
                                          .sum().reset_index(name='Total Segundos Esperados'))
                # Rename 'Nombre del empleado' to 'Nombre' in total_expected_summary to match summary_actual for merge
                total_expected_summary.rename(columns={'Nombre del empleado': 'Nombre'}, inplace=True)

                resumen_df = pd.merge(summary_actual, total_expected_summary, on=['ID Empleado', 'Nombre'], how='left')
                resumen_df['Total Segundos Esperados'] = resumen_df['Total Segundos Esperados'].fillna(0)
            else:
                resumen_df = summary_actual.copy()
                resumen_df['Total Segundos Esperados'] = 0.0

            resumen_df['Días del periodo'] = (resumen_df['last_day_worked'] - resumen_df['first_day_worked']).apply(lambda td: td.days + 1 if pd.notnull(td) else 0)
            resumen_df['Horas trabajadas'] = resumen_df['total_worked_time_sum'].apply(fmt_timedelta_to_str)
            resumen_df['Horas Trabajadas (Segundos)'] = resumen_df['total_worked_time_sum'].apply(lambda td: td.total_seconds() if pd.notnull(td) else 0)
            resumen_df['Total Segundos Esperados'] = pd.to_numeric(resumen_df['Total Segundos Esperados'], errors='coerce').fillna(0)
            resumen_df['Diferencia (Segundos)'] = resumen_df['Horas Trabajadas (Segundos)'] - resumen_df['Total Segundos Esperados']
            
            def format_seconds_to_hhmmss_with_sign(total_seconds):
                if pd.isna(total_seconds): return "00:00:00"
                sign = "-" if total_seconds < 0 else ""
                total_seconds = abs(total_seconds) 
                h = int(total_seconds // 3600)
                m = int((total_seconds % 3600) // 60)
                s = int(round(total_seconds % 60))
                return f"{sign}{h:02d}:{m:02d}:{s:02d}"
            
            resumen_df['Diferencia (HH:MM:SS)'] = resumen_df['Diferencia (Segundos)'].apply(format_seconds_to_hhmmss_with_sign)
            
            resumen_df.rename(columns={'days_actually_worked': 'Días trabajados'}, inplace=True, errors='ignore')
            
            resumen_df_cols_final = ['ID Empleado', 'Nombre', 'Días del periodo', 'Días trabajados', 
                                     'Horas trabajadas', 'Horas Trabajadas (Segundos)', 
                                     'Total Segundos Esperados', 'Diferencia (Segundos)', 'Diferencia (HH:MM:SS)']
            
            for col in resumen_df_cols_final: 
                if col not in resumen_df.columns:
                    default_val = ""
                    if col in ['Días del periodo', 'Días trabajados', 'Horas Trabajadas (Segundos)', 
                               'Total Segundos Esperados', 'Diferencia (Segundos)']:
                        default_val = 0.0 if '(Segundos)' in col or 'Total' in col else 0
                    resumen_df[col] = default_val
            
            resumen_df = resumen_df[resumen_df_cols_final]

            total_rows_for_detail_list = []
            for _, r_resumen_row in resumen_df.iterrows():
                emp_name_for_total = r_resumen_row['Nombre'] 
                sum_numeric_expected_seconds_for_emp = r_resumen_row['Total Segundos Esperados']

                # For 'Fecha' in Totals row, ensure it's a date object if it needs to be,
                # or format as string if just displaying days worked.
                # Here, it's int(r_resumen_row['Días trabajados']), so it's fine as is.
                dias_trabajados_val = int(r_resumen_row['Días trabajados']) if pd.notnull(r_resumen_row['Días trabajados']) else 0

                total_row_dict = {
                    'ID Empleado': r_resumen_row.get('ID Empleado', ""),
                    'Nombre del empleado': emp_name_for_total, 
                    'Turno': 'Totales',
                    'Fecha': dias_trabajados_val, # This is count of days, not a date for Totals row
                    'Día': '', 
                    'Horas esperadas': sum_numeric_expected_seconds_for_emp, 
                    'Horas totales': r_resumen_row['Horas trabajadas'] 
                }
                for c_col in checada_cols_in_report: total_row_dict[c_col] = ''
                total_rows_for_detail_list.append(total_row_dict)
            
            totals_to_append_df = pd.DataFrame(total_rows_for_detail_list)
            if not totals_to_append_df.empty:
                totals_to_append_df = totals_to_append_df.reindex(columns=display_report_df.columns) 

            final_detail_sheet_dfs = []
            if not display_report_df.empty: 
                # The 'Fecha' column in display_report_df should already be datetime.date objects
                # If not, ensure it here:
                # display_report_df['Fecha'] = pd.to_datetime(display_report_df['Fecha'], errors='coerce').dt.date
                for emp_name, daily_data_group in display_report_df.groupby('Nombre del empleado', sort=False):
                    final_detail_sheet_dfs.append(daily_data_group.sort_values('Fecha'))
                    emp_total_row = totals_to_append_df[totals_to_append_df['Nombre del empleado'] == emp_name]
                    if not emp_total_row.empty:
                        final_detail_sheet_dfs.append(emp_total_row)
            
            if not final_detail_sheet_dfs:
                final_detail_report_df = pd.DataFrame(columns=final_report_columns_ordered)
            else:
                final_detail_report_df = pd.concat(final_detail_sheet_dfs, ignore_index=True)

            # ***** CAMBIO AQUÍ: Formatear columna 'Fecha' ANTES de escribir en Excel *****
            if 'Fecha' in final_detail_report_df.columns:
                # Convert to datetime objects first (if they are not already, e.g. for 'Totales' rows)
                # then format to string. For 'Totales' rows, 'Fecha' is an int (days worked), so handle it.
                def format_fecha_col(val):
                    if isinstance(val, (datetime.date, pd.Timestamp)):
                        return val.strftime('%Y-%m-%d')
                    return val # Return as is if not a date (e.g. int for 'Totales' row)
                
                final_detail_report_df['Fecha'] = final_detail_report_df['Fecha'].apply(format_fecha_col)


            with pd.ExcelWriter(dst, engine='openpyxl') as writer:
                final_detail_report_df.to_excel(writer, index=False, sheet_name='Detalle')
                resumen_df.to_excel(writer, index=False, sheet_name='Resumen')

            self._format_excel(dst, resumen_df) 
            self._toggle_busy(False); self._set_status("Reporte generado exitosamente", "success")
            self._show_success_dialog(dst)

        except ValueError as ve:
            self._toggle_busy(False); self._set_status(f"Error de valor: {ve}", "error"); 
            messagebox.showerror("Error de Valor", f"Ocurrió un error con los datos:\n{ve}\n\n{traceback.format_exc()}")
            print(traceback.format_exc())
        except Exception as e:
            self._toggle_busy(False); self._set_status(f"Error: {e}", "error"); 
            messagebox.showerror("Error", f"Ocurrió un error inesperado:\n{e}\n\n{traceback.format_exc()}")
            print(traceback.format_exc())

    def _format_excel(self, path, resumen_data_df=None):
        wb=load_workbook(path)
        
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
        dark_orange_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid") 

        def _format_ws(ws, is_resumen_sheet=False, df_data_for_resumen=None):
            header_fill=PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font=Font(color="FFFFFF", bold=True)
            total_fill=PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") 
            bold_font=Font(bold=True)
            thin=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            if ws.max_row == 0: return

            col_names_map = {cell.value: get_column_letter(cell.column) for cell in ws[1]}

            for c_idx_plus_1 in range(1, ws.max_column + 1):
                cell = ws.cell(1, c_idx_plus_1)
                cell.fill = header_fill; cell.font = header_font; cell.border = thin
            
            if ws.title == 'Detalle':
                turno_col_letter = col_names_map.get('Turno')
                if turno_col_letter:
                    for r_idx_plus_1 in range(2, ws.max_row + 1): 
                        if ws[f"{turno_col_letter}{r_idx_plus_1}"].value == 'Totales':
                            for c_idx_plus_1_total in range(1, ws.max_column + 1):
                                cell_total = ws.cell(r_idx_plus_1, c_idx_plus_1_total)
                                cell_total.fill = total_fill; cell_total.font = bold_font; cell_total.border = thin
            
            if is_resumen_sheet and df_data_for_resumen is not None:
                diferencia_hhmmss_col_letter = col_names_map.get('Diferencia (HH:MM:SS)')
                diferencia_segundos_col_name = 'Diferencia (Segundos)'

                if diferencia_hhmmss_col_letter and diferencia_segundos_col_name in df_data_for_resumen.columns:
                    for r_idx, (_, row_data) in enumerate(df_data_for_resumen.iterrows()):
                        excel_row_num = r_idx + 2 
                        cell_to_format = ws[f"{diferencia_hhmmss_col_letter}{excel_row_num}"]
                        diferencia_sec_valor = row_data[diferencia_segundos_col_name]
                        
                        if pd.notnull(diferencia_sec_valor):
                            if diferencia_sec_valor < 0:
                                cell_to_format.fill = yellow_fill
                            elif diferencia_sec_valor > 0:
                                cell_to_format.fill = dark_orange_fill
            
            for col_letter_obj in ws.columns: 
                column_letter_str = col_letter_obj[0].column_letter
                max_len = 0
                for cell_in_col in col_letter_obj:
                    if cell_in_col.value is not None:
                        try: max_len = max(max_len, len(str(cell_in_col.value)))
                        except: pass 
                
                adjusted_width = max_len + 3 
                header_value = ws[f"{column_letter_str}1"].value 

                min_widths = {
                    'ID Empleado': 12, 'Nombre del empleado': 30, 'Nombre': 30, 'Turno': 10,
                    'Fecha': 12, 'Día': 12, 'Horas esperadas': 20, 'Horas totales': 15,   
                    'Horas trabajadas': 15, 'Horas Trabajadas (Segundos)': 22,
                    'Total Segundos Esperados': 22, 'Diferencia (Segundos)': 22, 
                    'Diferencia (HH:MM:SS)': 22, 
                    'Días del periodo': 18, 'Días trabajados': 18,
                }
                
                default_min_width = 10 if str(header_value).startswith('Checada') else 12
                adjusted_width = max(adjusted_width, min_widths.get(header_value, default_min_width))
                ws.column_dimensions[column_letter_str].width = adjusted_width
        
        for sheet_name_iter in wb.sheetnames:
            current_ws = wb[sheet_name_iter]
            if sheet_name_iter == 'Resumen':
                _format_ws(current_ws, is_resumen_sheet=True, df_data_for_resumen=resumen_data_df)
            else:
                _format_ws(current_ws)
        try:
            wb.save(path)
        except Exception as e_save:
            messagebox.showerror("Error al guardar", f"No se pudo guardar el archivo Excel:\n{e_save}\n\nAsegúrese de que el archivo no esté abierto.")
            print(f"Error al guardar Excel: {e_save}")


if __name__ == "__main__":
    root=Tk(); app=CheckadorApp(root); root.mainloop()